import streamlit as st
import pandas as pd
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Performance Motoristas | J&T",
    page_icon="📦",
    layout="wide"
)

# ─── Estilo ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #0f0f0f; }
    .block-container { padding-top: 1.5rem; }
    h1 { color: #D32F2F; font-weight: 700; }
    h3 { color: #EEEEEE; }
    .stDataFrame { border-radius: 8px; }
    .badge-ok   { background:#1b5e20; color:#a5d6a7; padding:3px 10px; border-radius:12px; font-size:0.85em; font-weight:600; }
    .badge-fail { background:#b71c1c; color:#ef9a9a; padding:3px 10px; border-radius:12px; font-size:0.85em; font-weight:600; }
    .badge-na   { background:#37474f; color:#b0bec5; padding:3px 10px; border-radius:12px; font-size:0.85em; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ─── Feriados Nacionais ───────────────────────────────────────────────────────
FERIADOS = {
    date(2025, 1, 1), date(2025, 4, 18), date(2025, 4, 21),
    date(2025, 5, 1), date(2025, 9, 7),  date(2025, 10, 12),
    date(2025, 11, 2), date(2025, 11, 15), date(2025, 12, 25),
    date(2026, 1, 1), date(2026, 4, 3),  date(2026, 4, 21),
    date(2026, 5, 1), date(2026, 9, 7),  date(2026, 10, 12),
    date(2026, 11, 2), date(2026, 11, 15), date(2026, 12, 25),
}

# ─── Cabeçalho ────────────────────────────────────────────────────────────────
col_logo, col_title = st.columns([1, 9])
with col_logo:
    st.markdown("## 📦")
with col_title:
    st.markdown("# Performance de Motoristas")
    st.caption("Análise de critérios de elegibilidade e pontuação | J&T Express – Belém/Ananindeua")

st.divider()

# ─── Upload ────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "📂 Carregar base (Carta de Porte de Entrega)",
    type=["xlsx"],
    help="Exporte do JMS: Carta de Porte de Entrega"
)

# ─── Parâmetros (sidebar) ─────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Parâmetros dos Critérios")
    min_pkgs_day = st.number_input("📦 Mín. pacotes por dia",     min_value=1, value=25, step=1)
    min_days     = st.number_input("📅 Mín. dias entregues",      min_value=1, value=24, step=1)
    min_sundays  = st.number_input("☀️ Mín. domingos entregados", min_value=1, value=3,  step=1)
    st.divider()
    st.markdown("**Legenda**")
    st.markdown("🟢 Critério atingido  \n🔴 Critério não atingido")
    st.divider()
    st.caption("Base pode estar incompleta — resultados refletem os dados disponíveis.")


# ─── Funções de pontuação ─────────────────────────────────────────────────────
def calc_pts_21h(pct):
    """≥90% → 200pts | ≥70% → proporcional | <70% → 0"""
    if pct >= 0.90:   return 200
    elif pct >= 0.70: return round((pct - 0.70) / 0.20 * 200)
    return 0

def calc_pts_dom_fer(n):
    """60+ → 40 | 40-59 → 30 | 10-39 → 20 | <10 → 0"""
    if n >= 60:   return 40
    elif n >= 40: return 30
    elif n >= 10: return 20
    return 0


# ─── Processamento ─────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Processando base...")
def processar(file_bytes, min_pkg, min_d, min_sun):
    df = pd.read_excel(io.BytesIO(file_bytes))

    df['dt_entrega']   = pd.to_datetime(df['Horário da entrega'], errors='coerce')
    df = df.dropna(subset=['dt_entrega', 'Responsável pela entrega'])
    df['data']         = df['dt_entrega'].dt.date
    df['hora']         = df['dt_entrega'].dt.hour
    df['dia_semana']   = df['dt_entrega'].dt.dayofweek   # 6 = domingo
    df['is_dom_fer']   = (df['dia_semana'] == 6) | df['data'].apply(lambda d: d in FERIADOS)
    df['entregue_21h'] = df['hora'] <= 21

    eleg_rows = []
    pts_rows  = []

    for motorista, grp in df.groupby('Responsável pela entrega'):
        pkgs_dia         = grp.groupby('data').size()
        min_dia          = int(pkgs_dia.min())
        total_pkgs       = int(pkgs_dia.sum())
        dias_trabalhados = int(pkgs_dia.shape[0])
        domingos         = grp[grp['dia_semana'] == 6]['data'].nunique()

        c1 = bool((pkgs_dia >= min_pkg).all())
        c2 = dias_trabalhados >= min_d
        c3 = domingos >= min_sun

        # ── Elegibilidade ──
        eleg_rows.append({
            'Motorista':            motorista,
            'Total Pacotes':        total_pkgs,
            'Dias Trabalhados':     dias_trabalhados,
            'Mín. Pacotes/Dia':     min_dia,
            f'≥{min_pkg} pkg/dia':  c1,
            f'≥{min_d} dias':       c2,
            f'≥{min_sun} domingos': c3,
            'Domingos Entregues':   int(domingos),
            'ELEGÍVEL':             c1 and c2 and c3,
        })

        # ── Pontuação ──
        dias_presenca = int((pkgs_dia >= 5).sum())
        pct_21h       = grp['entregue_21h'].mean()
        total_dom_fer = int(grp['is_dom_fer'].sum())
        pts_p         = dias_presenca * 10
        pts_21        = calc_pts_21h(pct_21h)
        pts_df        = calc_pts_dom_fer(total_dom_fer)

        pts_rows.append({
            'Motorista':           motorista,
            'Total Pacotes':       total_pkgs,
            'Dias c/ ≥5 Entregas': dias_presenca,
            'Pts Presença':        pts_p,
            'Entregas ≤21h':       int(grp['entregue_21h'].sum()),
            '% ≤21h':              round(pct_21h * 100, 1),
            'Pts ≤21h':            pts_21,
            'Pkg Dom+Feriados':    total_dom_fer,
            'Pts Dom+Feriados':    pts_df,
            'TOTAL PONTOS':        pts_p + pts_21 + pts_df,
            'ELEGÍVEL':            c1 and c2 and c3,
        })

    df_eleg = pd.DataFrame(eleg_rows).sort_values('Total Pacotes', ascending=False)
    df_pts  = (pd.DataFrame(pts_rows)
               .sort_values('TOTAL PONTOS', ascending=False)
               .reset_index(drop=True))
    df_pts.index += 1

    return df_eleg, df_pts, df


# ─── Gerador de Excel ──────────────────────────────────────────────────────────
def gerar_excel(df_eleg, df_pts, min_pkg, min_d, min_sun):
    wb = Workbook()

    RED   = "D32F2F"
    DARK  = "1a1a1a"
    MID   = "2d2d2d"
    WHITE = "FFFFFF"
    YELLOW= "F9A825"

    thin   = Side(style='thin', color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def escrever_aba(ws, df, titulo, subtitulo, bool_cols):
        nc = len(df.columns)
        ws.merge_cells(f"A1:{get_column_letter(nc)}1")
        ws["A1"] = titulo
        ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=WHITE)
        ws["A1"].fill      = PatternFill("solid", fgColor=RED)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{get_column_letter(nc)}2")
        ws["A2"] = subtitulo
        ws["A2"].font      = Font(name="Arial", italic=True, size=10, color="CCCCCC")
        ws["A2"].fill      = PatternFill("solid", fgColor=DARK)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.append([])

        headers = list(df.columns)
        ws.append(headers)
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
            cell.fill      = PatternFill("solid", fgColor=MID)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[4].height = 28

        for r_idx, row in enumerate(df.itertuples(index=False), start=5):
            elegivel = row[-1]
            row_bg   = "1e3a1e" if elegivel else "2d1a1a"
            for c_idx, (hdr, val) in enumerate(zip(headers, row), start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border    = border
                cell.font      = Font(name="Arial", size=9, color=WHITE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill      = PatternFill("solid", fgColor=row_bg)
                if hdr in bool_cols:
                    cell.value = "✔ SIM" if val else "✘ NÃO"
                    cell.font  = Font(name="Arial", size=9, bold=True,
                                      color="81C784" if val else "EF9A9A")
                elif hdr == "TOTAL PONTOS":
                    cell.value = val
                    cell.font  = Font(name="Arial", size=10, bold=True, color=YELLOW)
                else:
                    cell.value = val
                if hdr == "Motorista":
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for i, w in enumerate([38] + [14] * (nc - 1), 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        total      = len(df)
        total_eleg = int(df["ELEGÍVEL"].sum())
        sr = ws.max_row + 2
        ws.cell(row=sr, column=1, value="RESUMO").font = Font(
            name="Arial", bold=True, color=YELLOW, size=10)
        for offset, (label, val) in enumerate([
            ("Total de Motoristas", total),
            ("Elegíveis", total_eleg),
            ("Não Elegíveis", total - total_eleg),
        ], 1):
            ws.cell(row=sr+offset, column=1, value=label).font = Font(
                name="Arial", size=9, color="CCCCCC")
            ws.cell(row=sr+offset, column=2, value=val).font = Font(
                name="Arial", bold=True, size=9, color=WHITE)

    ws1 = wb.active
    ws1.title = "Elegibilidade"
    escrever_aba(
        ws1, df_eleg,
        "J&T Express – Elegibilidade de Motoristas",
        f"Critérios: ≥{min_pkg} pkg em todos os dias  |  ≥{min_d} dias no mês  |  ≥{min_sun} domingos",
        {f"≥{min_pkg} pkg/dia", f"≥{min_d} dias", f"≥{min_sun} domingos", "ELEGÍVEL"}
    )

    ws2 = wb.create_sheet("Pontuação")
    escrever_aba(
        ws2, df_pts.reset_index(drop=True),
        "J&T Express – Ranking de Pontuação",
        "Presença (+10/dia) · Entregas ≤21h (até 200pts) · Dom+Feriados (até 40pts)",
        {"ELEGÍVEL"}
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Main ──────────────────────────────────────────────────────────────────────
if uploaded:
    file_bytes = uploaded.read()
    df_eleg, df_pts, df_raw = processar(file_bytes, min_pkgs_day, min_days, min_sundays)

    # KPIs
    total     = len(df_eleg)
    elegiveis = int(df_eleg["ELEGÍVEL"].sum())
    nao_eleg  = total - elegiveis
    lider_pts = int(df_pts["TOTAL PONTOS"].iloc[0]) if len(df_pts) else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1: st.metric("🚴 Total Motoristas", total)
    with k2: st.metric("✅ Elegíveis",         elegiveis)
    with k3: st.metric("❌ Não Elegíveis",     nao_eleg)
    with k4: st.metric("📅 Dias na base",      df_raw['data'].nunique())
    with k5: st.metric("⭐ Maior Pontuação",   f"{lider_pts} pts")

    st.divider()

    tab1, tab2 = st.tabs(["📋 Elegibilidade", "⭐ Pontuação"])

    # ─── Tab Elegibilidade ────────────────────────────────────────────────────
    with tab1:
        filtro = st.radio(
            "Filtrar por:",
            ["Todos", "Apenas Elegíveis", "Apenas Não Elegíveis"],
            horizontal=True
        )
        df_show = (
            df_eleg[df_eleg["ELEGÍVEL"]]  if filtro == "Apenas Elegíveis"     else
            df_eleg[~df_eleg["ELEGÍVEL"]] if filtro == "Apenas Não Elegíveis" else
            df_eleg
        )

        c1_col = f"≥{min_pkgs_day} pkg/dia"
        c2_col = f"≥{min_days} dias"
        c3_col = f"≥{min_sundays} domingos"
        bool_cols_list = [c1_col, c2_col, c3_col, "ELEGÍVEL"]

        eleg_flag  = df_show["ELEGÍVEL"].values
        df_display = df_show.copy()
        for col in [c1_col, c2_col, c3_col]:
            df_display[col] = df_display[col].map({True: "✔ SIM", False: "✘ NÃO"})
        df_display["ELEGÍVEL"] = df_display["ELEGÍVEL"].map(
            {True: "✔ ELEGÍVEL", False: "✘ NÃO ELEGÍVEL"})

        def color_bool(val):
            if val in ("✔ SIM", "✔ ELEGÍVEL"):
                return "background-color:#1b5e20; color:#a5d6a7; font-weight:600"
            if val in ("✘ NÃO", "✘ NÃO ELEGÍVEL"):
                return "background-color:#7f0000; color:#ef9a9a; font-weight:600"
            return ""

        def color_row(row):
            idx  = df_display.index.get_loc(row.name)
            base = "background-color:#1e3a1e" if eleg_flag[idx] else "background-color:#2d1a1a"
            return [base] * len(row)

        st.dataframe(
            df_display.style
            .apply(color_row, axis=1)
            .map(color_bool, subset=bool_cols_list),
            use_container_width=True, height=560
        )

        with st.expander("🔍 Ver detalhamento por motorista (pacotes/dia)"):
            df_raw2 = pd.read_excel(io.BytesIO(file_bytes))
            df_raw2['dt_entrega'] = pd.to_datetime(df_raw2['Horário da entrega'], errors='coerce')
            df_raw2 = df_raw2.dropna(subset=['dt_entrega', 'Responsável pela entrega'])
            df_raw2['data'] = df_raw2['dt_entrega'].dt.date

            motorista_sel = st.selectbox(
                "Selecione o motorista:",
                sorted(df_raw2['Responsável pela entrega'].unique())
            )
            grp_sel = (
                df_raw2[df_raw2['Responsável pela entrega'] == motorista_sel]
                .groupby('data').size().reset_index(name='Pacotes')
            )
            grp_sel['Dia Semana'] = pd.to_datetime(grp_sel['data']).dt.day_name()
            grp_sel['Status']     = grp_sel['Pacotes'].apply(
                lambda x: "✔" if x >= min_pkgs_day else "✘")
            st.dataframe(grp_sel, use_container_width=True)

    # ─── Tab Pontuação ────────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📌 Regras de Pontuação")
        r1, r2, r3 = st.columns(3)
        with r1:
            st.info("**📍 Presença**\n\n"
                    "- ≥ 5 entregas no dia → **+10 pts**\n"
                    "- < 5 entregas → 0 pts\n"
                    "- Acumulado por dia válido")
        with r2:
            st.info("**🕘 Entregas até 21h**\n\n"
                    "- ≥ 90% das entregas → **+200 pts**\n"
                    "- ≥ 70% → proporcional\n"
                    "- < 70% → 0 pts")
        with r3:
            st.info("**📅 Domingos + Feriados**\n\n"
                    "- 60+ entregas → **+40 pts**\n"
                    "- 40 a 59 → **+30 pts**\n"
                    "- 10 a 39 → **+20 pts**\n"
                    "- < 10 → 0 pts")

        st.divider()
        st.markdown("### 🏆 Ranking de Pontuação")

        df_pts_display = df_pts.copy()
        df_pts_display["ELEGÍVEL"] = df_pts_display["ELEGÍVEL"].map(
            {True: "✔ SIM", False: "✘ NÃO"})
        df_pts_display["% ≤21h"] = df_pts_display["% ≤21h"].apply(lambda x: f"{x:.1f}%")

        def color_pts_bool(val):
            if val == "✔ SIM":
                return "background-color:#1b5e20; color:#a5d6a7; font-weight:600"
            if val == "✘ NÃO":
                return "background-color:#7f0000; color:#ef9a9a; font-weight:600"
            return ""

        def color_total(val):
            try:
                v = int(val)
                if v > 200: return "color:#F9A825; font-weight:700"
                if v > 100: return "color:#FFB74D; font-weight:600"
            except: pass
            return ""

        st.dataframe(
            df_pts_display.style
            .map(color_pts_bool, subset=["ELEGÍVEL"])
            .map(color_total,    subset=["TOTAL PONTOS"]),
            use_container_width=True, height=560
        )

    st.divider()

    excel_bytes = gerar_excel(df_eleg, df_pts, min_pkgs_day, min_days, min_sundays)
    st.download_button(
        label="📥 Baixar Relatório Excel (Elegibilidade + Pontuação)",
        data=excel_bytes,
        file_name="relatorio_performance_motoristas_jt.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )

else:
    st.info("👆 Faça o upload da base exportada do JMS para iniciar a análise.")
    st.markdown(f"""
    **Critérios de elegibilidade (configuráveis na barra lateral):**
    - 📦 **Mín. pacotes/dia** — pelo menos {min_pkgs_day} pacotes em *todos* os dias que trabalhou
    - 📅 **Mín. dias no mês** — pelo menos {min_days} dias distintos com entrega
    - ☀️ **Mín. domingos** — pelo menos {min_sundays} domingos distintos

    **Sistema de pontuação:**
    - 📍 Presença: +10 pts por dia com ≥ 5 entregas
    - 🕘 Entregas até 21h: até +200 pts (≥90% → 200, ≥70% → proporcional)
    - 📅 Domingos + Feriados: 60+→40pts / 40-59→30pts / 10-39→20pts
    """)